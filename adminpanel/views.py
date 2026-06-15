from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.contrib.contenttypes.models import ContentType

from live_webinars.models import LiveWebinar
from orders.models import OrderItem

from django.utils import timezone
import csv
from django.http import HttpResponse
import json
from integrations.models import ZoomMeeting
from .forms import InstructorForm
from django.contrib.admin.views.decorators import staff_member_required
import json
from integrations.services import link_existing_zoom_meeting

from django.shortcuts import render, redirect
from live_webinars.models import (
    LiveWebinar,
    
    WebinarPricing,
    Instructor,
    WebinarCategory,
    GlobalPricing
)

from recorded_webinars.models import (
    RecordedWebinar,
    
    RecordedWebinarPricing
)
from django.shortcuts import get_object_or_404

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from django.core.cache import cache

from datetime import timedelta
from django.db.models import Q, Count, Sum
from django.contrib.contenttypes.models import ContentType
from orders.models import Order
from live_webinars.models import LiveWebinar
from live_webinars.models import Instructor


@staff_member_required
def dashboard(request):
    """Optimized Admin Dashboard with Redis Cache"""

    cache_key = "admin_dashboard_data"

    # 🔥 Try cache first
    data = cache.get(cache_key)

    if not data:

        now = timezone.now()

        # ---------- DATE CALCULATIONS ----------
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_end = this_month_start - timezone.timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        # ---------- REVENUE ----------
        orders_paid = Order.objects.filter(status="PAID")

        this_month_revenue = orders_paid.filter(
            created_at__gte=this_month_start
        ).aggregate(total=Sum("items__price"))["total"] or 0

        last_month_revenue = orders_paid.filter(
            created_at__gte=last_month_start,
            created_at__lt=this_month_start
        ).aggregate(total=Sum("items__price"))["total"] or 0

        # ---------- TOTAL STATS ----------
        total_webinars = LiveWebinar.objects.count()
        total_instructors = Instructor.objects.count()
        total_orders = orders_paid.count()

        total_revenue = orders_paid.aggregate(
            total=Sum("items__price")
        )["total"] or 0

        # ---------- GROWTH ----------
        growth_months = []
        growth_orders = []

        for i in range(5, -1, -1):
            month_date = now - timezone.timedelta(days=30 * i)
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

            if month_start.month == 12:
                next_month_start = month_start.replace(year=month_start.year + 1, month=1)
            else:
                next_month_start = month_start.replace(month=month_start.month + 1)

            count = orders_paid.filter(
                created_at__gte=month_start,
                created_at__lt=next_month_start
            ).count()

            growth_months.append(month_start.strftime("%b"))
            growth_orders.append(count)

        # ---------- INSTRUCTOR ----------
        total_instructor_count = total_instructors

        instructor_months = []
        instructor_counts = []

        for i in range(5, -1, -1):
            month_date = now - timezone.timedelta(days=30 * i)
            month_start = month_date.replace(day=1)

            instructor_months.append(month_start.strftime("%b"))
            instructor_counts.append(total_instructor_count)

        # ---------- FINAL DATA ----------
        data = {
            "total_webinars": total_webinars,
            "total_instructors": total_instructors,
            "total_orders": total_orders,
            "total_revenue": round(total_revenue, 2),

            "this_month_revenue": round(this_month_revenue, 2),
            "last_month_revenue": round(last_month_revenue, 2),

            "growth_months": growth_months,
            "growth_orders": growth_orders,

            "instructor_months": instructor_months,
            "instructor_counts": instructor_counts,
        }

        # 🔥 Cache for 5 mins
        cache.set(cache_key, data, 300)

    return render(request, "adminpanel/dashboard.html", data)


@staff_member_required
def webinar_management(request):

    # -----------------------
    # 🔥 CACHE KEY
    # -----------------------
    cache_key = f"admin_webinars_{request.GET.urlencode()}"
    cached_data = cache.get(cache_key)

    if cached_data:
        return render(request, "adminpanel/webinar_list.html", cached_data)

    # -----------------------
    # FETCH DATA (OPTIMIZED)
    # -----------------------
    live_webinars = LiveWebinar.objects.select_related("instructor", "category").only(
        "id", "title", "start_datetime", "duration_minutes",
        "instructor__name", "category__name"
    )

    recorded_webinars = RecordedWebinar.objects.select_related("instructor", "category").only(
        "id", "title", "created_at", "duration_minutes",
        "instructor__name", "category__name"
    ).filter(is_published=True)

    # -----------------------
    # SEARCH
    # -----------------------
    search = request.GET.get("search", "").strip()
    if search:
        live_webinars = live_webinars.filter(title__icontains=search)
        recorded_webinars = recorded_webinars.filter(title__icontains=search)

    # -----------------------
    # STATUS FILTER (NO LOOP 🚀)
    # -----------------------
    status_filter = request.GET.get("status", "all")
    now = timezone.now()

    if status_filter == "RECORDED":
        live_webinars = LiveWebinar.objects.none()

    elif status_filter == "UPCOMING":
        live_webinars = live_webinars.filter(start_datetime__gt=now)
        recorded_webinars = RecordedWebinar.objects.none()

    elif status_filter == "LIVE":
        live_webinars = live_webinars.filter(start_datetime__lte=now)
        recorded_webinars = RecordedWebinar.objects.none()

    elif status_filter == "ENDED":
        live_webinars = live_webinars.filter(
            start_datetime__lt=now - timedelta(minutes=120)
        )
        recorded_webinars = RecordedWebinar.objects.none()

    # -----------------------
    # 🔥 STATS CACHE
    # -----------------------
    stats_cache = cache.get("webinar_stats")

    if not stats_cache:

        live_ct = ContentType.objects.get_for_model(LiveWebinar)
        recorded_ct = ContentType.objects.get_for_model(RecordedWebinar)

        live_stats = OrderItem.objects.filter(
            content_type=live_ct,
            order__status="PAID"
        ).values("object_id").annotate(
            enrolled=Count("id"),
            revenue=Sum("price")
        )

        recorded_stats = OrderItem.objects.filter(
            content_type=recorded_ct,
            order__status="PAID"
        ).values("object_id").annotate(
            enrolled=Count("id"),
            revenue=Sum("price")
        )

        stats_cache = (live_stats, recorded_stats)
        cache.set("webinar_stats", stats_cache, 300)

    else:
        live_stats, recorded_stats = stats_cache

    # -----------------------
    # MAP
    # -----------------------
    live_stats_map = {item["object_id"]: item for item in live_stats}
    recorded_stats_map = {item["object_id"]: item for item in recorded_stats}

    # -----------------------
    # BUILD DATA
    # -----------------------
    webinar_data = []

    for webinar in live_webinars:
        stat = live_stats_map.get(webinar.id, {})
        webinar_data.append({
            "obj": webinar,
            "type": "live",
            "dynamic_status": "LIVE",
            "enrolled": stat.get("enrolled", 0),
            "revenue": stat.get("revenue", 0)
        })

    for webinar in recorded_webinars:
        stat = recorded_stats_map.get(webinar.id, {})
        webinar_data.append({
            "obj": webinar,
            "type": "recorded",
            "dynamic_status": "RECORDED",
            "enrolled": stat.get("enrolled", 0),
            "revenue": stat.get("revenue", 0)
        })

    # -----------------------
    # SORT
    # -----------------------
    webinar_data.sort(
        key=lambda x: x["obj"].start_datetime if x["type"] == "live" else x["obj"].created_at,
        reverse=True
    )

    # -----------------------
    # PAGINATION
    # -----------------------
    paginator = Paginator(webinar_data, 15)
    page = request.GET.get("page")
    webinars_page = paginator.get_page(page)

    context = {
        "webinars": webinars_page,
        "count": len(webinar_data),
        "page_obj": webinars_page,
        "search": search,
        "status_filter": status_filter
    }

    # -----------------------
    # SAVE CACHE
    # -----------------------
    cache.set(cache_key, context, 120)

    return render(request, "adminpanel/webinar_list.html", context)



from django.shortcuts import render, redirect
from live_webinars.models import (
    LiveWebinar,
    WebinarPricing,
    WebinarCategory,
    Instructor,
)
from recorded_webinars.models import (
    RecordedWebinar,
    RecordedWebinarPricing,
)
from live_webinars.models import GlobalPricing


@staff_member_required
def create_webinar(request):

    instructors = Instructor.objects.all()
    categories = WebinarCategory.objects.all()
    pricing = GlobalPricing.get_pricing()
    
    # Calculate total webinars (both live and recorded)
    total_live = LiveWebinar.objects.count()
    total_recorded = RecordedWebinar.objects.filter(is_published=True).count()
    total_webinars = total_live + total_recorded

    if request.method == "POST":

        webinar_type = request.POST.get("type")

        title = request.POST.get("title")
        category_id = request.POST.get("category")
        instructor_id = request.POST.get("instructor")

        duration = request.POST.get("duration")
        duration = int(duration) if duration else 60

        description = request.POST.get("description")

        category = WebinarCategory.objects.get(id=category_id)
        instructor = Instructor.objects.get(id=instructor_id)

        # =========================================
        # LIVE WEBINAR
        # =========================================

        if webinar_type == "live":

            from datetime import datetime

            start_datetime_str = request.POST.get("start_datetime")

            if not start_datetime_str:
                return redirect("adminpanel:webinar_management")

            start_datetime = datetime.fromisoformat(start_datetime_str)

            webinar = LiveWebinar.objects.create(
                title=title,
                category=category,
                instructor=instructor,
                start_datetime=start_datetime,
                duration_minutes=duration,
                description=description
            )

            # Zoom meeting
            zoom_json = request.POST.get("selected_zoom_meeting")

            if zoom_json:
                meeting_data = json.loads(zoom_json)
                link_existing_zoom_meeting(webinar, meeting_data)

            # Content
            

            # Pricing
            live_single = request.POST.get("live_single_price")

            if live_single:
                WebinarPricing.objects.create(
                    webinar=webinar,
                    live_single_price=live_single,
                    live_multi_price=request.POST.get("live_multi_price"),
                    recorded_single_price=request.POST.get("recorded_single_price"),
                    recorded_multi_price=request.POST.get("recorded_multi_price"),
                    combo_single_price=request.POST.get("combo_single_price"),
                    combo_multi_price=request.POST.get("combo_multi_price"),
                )

        # =========================================
        # RECORDED WEBINAR
        # =========================================

        else:

            recording_json = request.POST.get("selected_zoom_recording")

            recording_link = None

            if recording_json:
                rec = json.loads(recording_json)
                recording_link = rec["play_url"]

            webinar = RecordedWebinar.objects.create(
                title=title,
                category=category,
                instructor=instructor,
                duration_minutes=duration,
                zoom_recording_link=recording_link,
                is_published=True,
                description=description
            )

   

            single_price = request.POST.get("recorded_single_price_only")

            if single_price:
                RecordedWebinarPricing.objects.create(
                    webinar=webinar,
                    single_price=single_price,
                    multi_user_price=request.POST.get("recorded_multi_price_only"),
                )

        return redirect("adminpanel:webinar_management")

    return render(
        request,
        "adminpanel/webinar_create.html",
        {
            "instructors": instructors,
            "categories": categories,
            "pricing": pricing,
            "total_webinars": total_webinars,
        },
    )

from django.http import JsonResponse
from live_webinars.models import LiveWebinar


@staff_member_required
def webinar_status_api(request):

    webinars = LiveWebinar.objects.all()

    data = []

    for w in webinars:
        data.append({
            "id": w.id,
            "status": w.status
        })

    return JsonResponse({"webinars": data})


@staff_member_required
def delete_webinar(request, id):

    webinar = get_object_or_404(LiveWebinar, id=id)

    webinar.delete()

    return redirect("adminpanel:webinar_management")


@staff_member_required
def edit_webinar(request, id):

    webinar = get_object_or_404(LiveWebinar, id=id)

    instructors = Instructor.objects.all()
    categories = WebinarCategory.objects.all()
    pricing = GlobalPricing.get_pricing()

    if request.method == "POST":

        webinar.title = request.POST.get("title")
        webinar.category_id = request.POST.get("category")
        webinar.instructor_id = request.POST.get("instructor")

        duration = request.POST.get("duration")
        webinar.duration_minutes = int(duration) if duration else 60

        # webinar.zoom_link = request.POST.get("zoom_link")

        from datetime import datetime
        start_datetime_str = request.POST.get("start_datetime")

        if start_datetime_str:
            webinar.start_datetime = datetime.fromisoformat(start_datetime_str)
        webinar.description = request.POST.get("description")
        webinar.save()

        # Update content sections

        
        selected_zoom = request.POST.get("selected_zoom_meeting")

        if selected_zoom:

            meeting_data = json.loads(selected_zoom)

            # delete old meeting
            if hasattr(webinar, "zoom_meeting"):
                webinar.zoom_meeting.delete()

            link_existing_zoom_meeting(webinar, meeting_data)

        return redirect("adminpanel:webinar_management")

    

    return render(
        request,
        "adminpanel/webinar_edit.html",
        {
            "webinar": webinar,
            
            "instructors": instructors,
            "categories": categories,
            "pricing": pricing,
        },
    )


@staff_member_required
def instructor_management(request):

    instructors = Instructor.objects.all().order_by("name")

    # SEARCH
    search = request.GET.get("search")

    if search:
        instructors = instructors.filter(name__icontains=search)

    paginator = Paginator(instructors, 15)
    page = request.GET.get("page")
    instructors_page = paginator.get_page(page)

    context = {
        "instructors": instructors_page,
        "count": paginator.count,
        "search": search,
        "page_obj": instructors_page
    }

    return render(
        request,
        "adminpanel/instructor_list.html",
        context
    )


@staff_member_required
def create_instructor(request):

    if request.method == "POST":

        Instructor.objects.create(

            name=request.POST.get("name"),

            designation=request.POST.get("designation"),

            organization=request.POST.get("organization"),

            email=request.POST.get("email"),

            country=request.POST.get("country"),

            phone_number=request.POST.get("phone_number"),

            bio=request.POST.get("bio"),

            photo=request.FILES.get("photo")

        )

        return redirect("adminpanel:instructor_management")

    return render(request,"adminpanel/instructor_create.html")

from django.shortcuts import get_object_or_404


@staff_member_required
def edit_instructor(request, id):

    instructor = get_object_or_404(Instructor, id=id)

    if request.method == "POST":

        instructor.name = request.POST.get("name")
        instructor.designation = request.POST.get("designation")
        instructor.organization = request.POST.get("organization")
        instructor.email = request.POST.get("email")
        instructor.country = request.POST.get("country")
        instructor.phone_number = request.POST.get("phone_number")
        instructor.bio = request.POST.get("bio")

        if request.FILES.get("photo"):
            instructor.photo = request.FILES.get("photo")

        instructor.save()

        return redirect("adminpanel:instructor_management")

    return render(
        request,
        "adminpanel/instructor_edit.html",
        {"instructor": instructor}
    )


@staff_member_required
def delete_instructor(request, id):

    instructor = get_object_or_404(Instructor, id=id)

    instructor.delete()

    return redirect("adminpanel:instructor_management")


@staff_member_required
def instructor_detail(request, id):

    instructor = get_object_or_404(Instructor, id=id)

    return render(
        request,
        "adminpanel/instructor_detail.html",
        {"instructor": instructor}
    )

from orders.models import Order
from django.db.models import Prefetch


@staff_member_required
def payment_history(request):

    # ✅ GET ALL ORDERS (IMPORTANT FIX)
    orders = Order.objects.all()\
        .prefetch_related("items")\
        .select_related("user")\
        .order_by("-created_at")

    # ========== DATE RANGE FILTER ==========
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if from_date and to_date:
        from datetime import datetime
        try:
            from_date_obj = datetime.strptime(from_date, "%d-%m-%Y")
            to_date_obj = datetime.strptime(to_date, "%d-%m-%Y")

            from_date_obj = timezone.make_aware(
                from_date_obj.replace(hour=0, minute=0, second=0)
            )
            to_date_obj = timezone.make_aware(
                to_date_obj.replace(hour=23, minute=59, second=59)
            )

            orders = orders.filter(created_at__range=[from_date_obj, to_date_obj])

        except ValueError:
            pass

    # ========== BUILD PAYMENT DATA ==========
    payment_data = []

    for order in orders:
        for item in order.items.all():

            payment_data.append({
                "date": order.created_at,
                "name": order.user.get_full_name() or order.user.email,
                "email": order.user.email,

                # ✅ FIXED (NO .title mistake)
                "webinar": item.webinar_title or "Unknown",

                "amount": item.price,
                "variant": item.variant,

                # ✅ IMPORTANT (SHOW STATUS)
                "status": order.status,
                "order_number": order.order_number
            })

    # ========== REVENUE STATS (ONLY PAID) ==========
    now = timezone.now()

    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    week_start = now - timezone.timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)

    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    paid_orders = Order.objects.filter(status="PAID")

    # Today
    today_orders = paid_orders.filter(created_at__gte=today_start)
    today_revenue = today_orders.aggregate(total=Sum("items__price"))["total"] or 0
    today_count = today_orders.count()

    # Week
    week_orders = paid_orders.filter(created_at__gte=week_start)
    week_revenue = week_orders.aggregate(total=Sum("items__price"))["total"] or 0
    week_count = week_orders.count()

    # Month
    month_orders = paid_orders.filter(created_at__gte=month_start)
    month_revenue = month_orders.aggregate(total=Sum("items__price"))["total"] or 0
    month_count = month_orders.count()

    # Total
    total_revenue = paid_orders.aggregate(total=Sum("items__price"))["total"] or 0
    total_count = paid_orders.count()

    return render(request, "adminpanel/payment_history.html", {
        "payments": payment_data,

        "today_revenue": round(today_revenue, 2),
        "today_count": today_count,

        "week_revenue": round(week_revenue, 2),
        "week_count": week_count,

        "month_revenue": round(month_revenue, 2),
        "month_count": month_count,

        "total_revenue": round(total_revenue, 2),
        "total_count": total_count,

        "from_date": from_date,
        "to_date": to_date,
    })

from orders.models import Order
from django.shortcuts import render, get_object_or_404

@staff_member_required
def payment_detail(request, order_number):

    order = get_object_or_404(
        Order.objects.select_related("user").prefetch_related("items"),
        order_number=order_number
    )

    items = []

    for item in order.items.all():

        items.append({
            "title": item.webinar_title or "Unknown Webinar",   # ✅ FIX
            "price": item.price,
            "variant": item.variant,
            "date": order.created_at,

            # ✅ BONUS (since you store it)
            "instructor": item.instructor_name,
            "category": item.category_name
        })

    return render(
        request,
        "adminpanel/payment_detail.html",
        {
            "order": order,
            "items": items
        }
    )

from orders.email_utils import send_invoice_email
from django.shortcuts import redirect


@staff_member_required
def resend_invoice(request, order_number):

    order = get_object_or_404(Order, order_number=order_number)

    send_invoice_email(order)

    return redirect("adminpanel:payment_detail", order_number=order_number)


# ========== EXPORT FUNCTIONS ==========

@staff_member_required
def export_payments_csv(request):
    orders = Order.objects.filter(
        status="PAID"
    ).prefetch_related("items").select_related("user").order_by("-created_at")

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if from_date and to_date and from_date != "None" and to_date != "None":
        from datetime import datetime
        try:
            from_date_obj = datetime.strptime(from_date, "%d-%m-%Y")
            to_date_obj = datetime.strptime(to_date, "%d-%m-%Y")

            from_date_obj = timezone.make_aware(from_date_obj.replace(hour=0, minute=0, second=0))
            to_date_obj = timezone.make_aware(to_date_obj.replace(hour=23, minute=59, second=59))

            orders = orders.filter(created_at__range=[from_date_obj, to_date_obj])
        except:
            pass

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="payment_history.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Order Number", "Date", "Customer Name", "Email",
        "Webinar", "Instructor", "Category",
        "Amount", "Access Type", "Status"
    ])

    for order in orders:
        for item in order.items.all():
            writer.writerow([
                order.order_number,
                order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                order.user.get_full_name() or order.user.email,
                order.user.email,
                item.webinar_title or "Unknown",
                item.instructor_name or "N/A",
                item.category_name or "N/A",
                item.price,
                item.variant,
                order.status
            ])

    return response


@staff_member_required
def export_payments_excel(request):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    orders = Order.objects.filter(
        status="PAID"
    ).prefetch_related("items").select_related("user").order_by("-created_at")

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if from_date and to_date and from_date != "None" and to_date != "None":
        from datetime import datetime
        try:
            from_date_obj = datetime.strptime(from_date, "%d-%m-%Y")
            to_date_obj = datetime.strptime(to_date, "%d-%m-%Y")

            from_date_obj = timezone.make_aware(from_date_obj.replace(hour=0, minute=0, second=0))
            to_date_obj = timezone.make_aware(to_date_obj.replace(hour=23, minute=59, second=59))

            orders = orders.filter(created_at__range=[from_date_obj, to_date_obj])
        except:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = [
        "Order Number", "Date", "Customer Name", "Email",
        "Webinar", "Instructor", "Category",
        "Amount", "Access Type", "Status"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    row_num = 2

    for order in orders:
        for item in order.items.all():
            ws.append([
                order.order_number,
                order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                order.user.get_full_name() or order.user.email,
                order.user.email,
                item.webinar_title or "Unknown",
                item.instructor_name or "N/A",
                item.category_name or "N/A",
                float(item.price),
                item.variant,
                order.status
            ])
            row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="payment_history.xlsx"'

    wb.save(response)
    return response

@staff_member_required
def subscription_management(request):
    """Professional Subscription Management Dashboard"""
    from subscriptions.models import SubscriptionPlan, UserSubscription, SubscriptionPayment
    from django.db.models import Count, Sum
    
    # Get all subscription plans
    plans = SubscriptionPlan.objects.all()
    
    # Get all user subscriptions with search and filter
    subscriptions = UserSubscription.objects.select_related('user', 'plan').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        subscriptions = subscriptions.filter(
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(plan__name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)
    
    # Filter by plan
    plan_filter = request.GET.get('plan', '')
    if plan_filter:
        subscriptions = subscriptions.filter(plan_id=plan_filter)
    
    # Pagination
    paginator = Paginator(subscriptions, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_subscriptions = UserSubscription.objects.count()
    active_subscriptions = UserSubscription.objects.filter(status="ACTIVE").count()
    expired_subscriptions = UserSubscription.objects.filter(status="EXPIRED").count()
    
    # Revenue from subscriptions
    payments = SubscriptionPayment.objects.filter(status="PAID").aggregate(
        total=Sum('plan__price')
    )
    total_revenue = payments['total'] or 0
    
    # Subscriptions by plan
    plan_stats = SubscriptionPlan.objects.annotate(
        subscriber_count=Count('usersubscription')
    )
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'plan_filter': plan_filter,
        'plans': plans,
        'plan_stats': plan_stats,
        'total_subscriptions': total_subscriptions,
        'active_subscriptions': active_subscriptions,
        'expired_subscriptions': expired_subscriptions,
        'total_revenue': total_revenue,
    }
    
    return render(request, 'adminpanel/subscription_management.html', context)


# ========== USER MANAGEMENT ==========

from accounts.models import CustomUser
from django.db.models import Q

@staff_member_required
def user_management(request):
    """User Management - List all users with search and filters"""
    
    users = CustomUser.objects.all().order_by('-created_at')
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(company__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    elif status_filter == 'verified':
        users = users.filter(is_verified=True)
    elif status_filter == 'unverified':
        users = users.filter(is_verified=False)
    elif status_filter == 'staff':
        users = users.filter(is_staff=True)
    
    # Get statistics for each user
    user_data = []
    for user in users:
        orders = Order.objects.filter(user=user, status='PAID')
        total_spent = orders.aggregate(total=Sum('items__price'))['total'] or 0
        total_orders = orders.count()
        
        user_data.append({
            'user': user,
            'total_orders': total_orders,
            'total_spent': round(total_spent, 2)
        })
    
    # Pagination
    paginator = Paginator(user_data, 20)
    page = request.GET.get('page')
    users_page = paginator.get_page(page)
    
    # Overall statistics
    total_users = CustomUser.objects.count()
    active_users = CustomUser.objects.filter(is_active=True).count()
    verified_users = CustomUser.objects.filter(is_verified=True).count()
    staff_users = CustomUser.objects.filter(is_staff=True).count()
    
    context = {
        'users': users_page,
        'page_obj': users_page,
        'search': search,
        'status_filter': status_filter,
        'total_users': total_users,
        'active_users': active_users,
        'verified_users': verified_users,
        'staff_users': staff_users,
    }
    
    return render(request, 'adminpanel/user_list.html', context)


def user_detail(request, id):
    """User Detail - View detailed information about a user"""
    
    user = get_object_or_404(CustomUser, id=id)
    
    # Get user's orders
    orders = Order.objects.filter(user=user).prefetch_related("items").order_by('-created_at')
    
    # Calculate statistics
    total_orders = orders.filter(status='PAID').count()
    total_spent = orders.filter(status='PAID').aggregate(
        total=Sum('items__price')
    )['total'] or 0
    
    # Get enrolled webinars (SNAPSHOT BASED)
    enrolled_webinars = []

    for order in orders.filter(status='PAID'):
        for item in order.items.all():

            if item.webinar_title:   # ✅ FIX

                enrolled_webinars.append({
                    'title': item.webinar_title,
                    'variant': item.variant,
                    'price': item.price,
                    'instructor': item.instructor_name or "N/A",
                    'category': item.category_name or "N/A",
                    'date': order.created_at
                })
    
    context = {
        'user_obj': user,
        'orders': orders[:10],
        'total_orders': total_orders,
        'total_spent': round(total_spent, 2),
        'enrolled_webinars': enrolled_webinars,
    }
    
    return render(request, 'adminpanel/user_detail.html', context)

@staff_member_required
def user_edit(request, id):
    """User Edit - Update user information"""
    
    user = get_object_or_404(CustomUser, id=id)
    
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.company = request.POST.get('company', '')
        user.phone_number = request.POST.get('phone_number', '')
        user.country = request.POST.get('country', '')
        
        # Update status fields
        user.is_active = request.POST.get('is_active') == 'on'
        user.is_verified = request.POST.get('is_verified') == 'on'
        user.is_staff = request.POST.get('is_staff') == 'on'
        
        user.save()
        
        return redirect('adminpanel:user_detail', id=user.id)
    
    context = {
        'user_obj': user,
    }
    
    return render(request, 'adminpanel/user_edit.html', context)

@staff_member_required
def user_delete(request, id):
    """User Delete - Delete a user"""
    
    user = get_object_or_404(CustomUser, id=id)
    user.delete()
    
    return redirect('adminpanel:user_management')


from recorded_webinars.models import RecordedWebinar
from django.shortcuts import get_object_or_404, render, redirect

@staff_member_required
def edit_recorded_webinar(request, id):

    webinar = get_object_or_404(RecordedWebinar, id=id)

    instructors = Instructor.objects.all()
    categories = WebinarCategory.objects.all()

    if request.method == "POST":

        webinar.title = request.POST.get("title")
        webinar.category_id = request.POST.get("category")
        webinar.instructor_id = request.POST.get("instructor")

        webinar.duration_minutes = int(request.POST.get("duration") or 60)

        webinar.description = request.POST.get("description")

        # recording update
        recording_json = request.POST.get("selected_zoom_recording")

        if recording_json:
            import json
            rec = json.loads(recording_json)
            webinar.zoom_recording_link = rec.get("play_url")

        webinar.save()

        return redirect("adminpanel:webinar_management")

    return render(request, "adminpanel/recorded_webinar_edit.html", {
        "webinar": webinar,
        "instructors": instructors,
        "categories": categories,
    })

@staff_member_required
def duplicate_webinar(request, id):

    webinar = get_object_or_404(LiveWebinar, id=id)

    # =========================
    # CREATE NEW WEBINAR
    # =========================

    duplicated = LiveWebinar.objects.create(
        title=f"{webinar.title} (Copy)",
        category=webinar.category,
        instructor=webinar.instructor,
        start_datetime=webinar.start_datetime,
        duration_minutes=webinar.duration_minutes,
        description=webinar.description,
    )

    # =========================
    # DUPLICATE PRICING
    # =========================

    try:
        old_price = WebinarPricing.objects.get(webinar=webinar)

        WebinarPricing.objects.create(
            webinar=duplicated,

            live_single_price=old_price.live_single_price,
            live_multi_price=old_price.live_multi_price,

            recorded_single_price=old_price.recorded_single_price,
            recorded_multi_price=old_price.recorded_multi_price,

            combo_single_price=old_price.combo_single_price,
            combo_multi_price=old_price.combo_multi_price,
        )

    except WebinarPricing.DoesNotExist:
        pass

    # =========================
    # DUPLICATE ZOOM MEETING
    # =========================

    if hasattr(webinar, "zoom_meeting"):

        old_zoom = webinar.zoom_meeting

        ZoomMeeting.objects.create(
            webinar=duplicated,

            zoom_meeting_id=old_zoom.zoom_meeting_id,
            uuid=old_zoom.uuid,

            topic=old_zoom.topic,

            join_url=old_zoom.join_url,
            start_url=old_zoom.start_url,

            password=old_zoom.password,

            start_time=old_zoom.start_time,
            duration=old_zoom.duration,

            timezone=old_zoom.timezone,
        )

    return redirect("adminpanel:edit_webinar", id=duplicated.id)

@staff_member_required
def duplicate_recorded_webinar(request, id):

    webinar = get_object_or_404(RecordedWebinar, id=id)

    duplicated = RecordedWebinar.objects.create(
        title=f"{webinar.title} (Copy)",
        category=webinar.category,
        instructor=webinar.instructor,
        duration_minutes=webinar.duration_minutes,
        description=webinar.description,
        zoom_recording_link=webinar.zoom_recording_link,
        is_published=webinar.is_published,
    )

    try:
        old_price = RecordedWebinarPricing.objects.get(webinar=webinar)

        RecordedWebinarPricing.objects.create(
            webinar=duplicated,
            single_price=old_price.single_price,
            multi_user_price=old_price.multi_user_price,
        )

    except RecordedWebinarPricing.DoesNotExist:
        pass

    return redirect(
        "adminpanel:edit_recorded_webinar",
        id=duplicated.id
    )